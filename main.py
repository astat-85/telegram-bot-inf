#!/usr/bin/env python3
"""
Telegram Bot для сбора игровых данных
АДАПТИРОВАНО ДЛЯ BOTHOST.RU
ПОЛНОСТЬЮ ИСПРАВЛЕННАЯ ВЕРСИЯ
"""
import sqlite3
import csv
import asyncio
import logging
import logging.handlers
import os
import sys
import threading
import shutil
import traceback
import json
import time
import re
import html
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional, Dict, Any, List, Union
from functools import wraps
from threading import RLock
from collections import defaultdict
from config import get_field_label, get_field_name_only, get_visible_fields

# ========== НАСТРОЙКИ ДЛЯ СЕТИ ==========
import aiohttp
connector = None  # Будет создан внутри async функции
timeout = aiohttp.ClientTimeout(total=30, connect=10, sock_connect=10, sock_read=10)

# ========== ПУТИ ==========
BASE_DIR = Path(__file__).parent
print(f"📁 Директория: {BASE_DIR}")

# ========== ПАПКА ДЛЯ ДАННЫХ ==========
DATA_DIR = BASE_DIR / "data"
DATA_DIR.mkdir(exist_ok=True, parents=True)
print(f"📁 Папка данных: {DATA_DIR}")

# ========== ПЕРЕМЕННЫЕ ОКРУЖЕНИЯ ==========
BOT_TOKEN = os.getenv("BOT_TOKEN", "").strip()
ADMIN_IDS_STR = os.getenv("ADMIN_IDS", "").strip()
TARGET_CHAT_ID = os.getenv("TARGET_CHAT_ID", "").strip()
TARGET_TOPIC_ID = os.getenv("TARGET_TOPIC_ID", "").strip()
DB_NAME = os.getenv("DB_NAME", str(DATA_DIR / "users_data.db"))

# ========== ВАЛИДАЦИЯ ТОКЕНА ==========
if not BOT_TOKEN or not re.match(r'^\d+:[\w-]+$', BOT_TOKEN):
    print("=" * 60)
    print("❌ ОШИБКА: BOT_TOKEN не установлен или неверный формат!")
    print("Добавьте в переменные окружения на Bothost.ru:")
    print("BOT_TOKEN = ваш_токен_бота")
    print("=" * 60)
    sys.exit(1)

# ========== ПАРСИНГ ID ==========
ADMIN_IDS = [int(x.strip()) for x in ADMIN_IDS_STR.split(',') if x.strip().isdigit()]
try:
    TARGET_CHAT_ID = int(TARGET_CHAT_ID) if TARGET_CHAT_ID else None
except ValueError:
    print(f"❌ ОШИБКА: TARGET_CHAT_ID должен быть числом: '{TARGET_CHAT_ID}'")
    TARGET_CHAT_ID = None

USE_TOPIC = False
if TARGET_TOPIC_ID and TARGET_TOPIC_ID.strip() not in ("", "0", "None", "none", "null"):
    try:
        TARGET_TOPIC_ID = int(TARGET_TOPIC_ID)
        USE_TOPIC = True
        print(f"✅ Тема: {TARGET_TOPIC_ID}")
    except ValueError:
        print(f"⚠️ Неверный TARGET_TOPIC_ID: '{TARGET_TOPIC_ID}'")

# ========== ДИРЕКТОРИИ ==========
EXPORT_DIR = DATA_DIR / "exports"
BACKUP_DIR = DATA_DIR / "backups"
LOGS_DIR = DATA_DIR / "logs"
for dir_path in [EXPORT_DIR, BACKUP_DIR, LOGS_DIR]:
    dir_path.mkdir(exist_ok=True, parents=True)

# ========== ЛОГИРОВАНИЕ ==========
log_handler = logging.handlers.RotatingFileHandler(
    LOGS_DIR / 'bot.log',
    maxBytes=10*1024*1024,
    backupCount=5,
    encoding='utf-8'
)
log_handler.setFormatter(logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s'))
logging.basicConfig(
    level=logging.INFO,
    handlers=[log_handler, logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger(__name__)

# ========== AIOGRAM ==========
try:
    from aiogram import Bot, Dispatcher, Router, F
    from aiogram.fsm.context import FSMContext
    from aiogram.fsm.state import State, StatesGroup
    from aiogram.fsm.storage.memory import MemoryStorage
    from aiogram.filters import Command
    from aiogram.types import (
        ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup,
        InlineKeyboardButton, Message, CallbackQuery, FSInputFile, ChatMemberUpdated
    )
    from aiogram.exceptions import TelegramBadRequest
    from aiogram.enums import ParseMode
    from handlers import profile
    from database.profile_db import ProfileDB
    from cities.city_db import CityDatabase
    from keyboards.profile import (
        get_profile_menu_keyboard, get_edit_profile_keyboard,
        get_city_choice_keyboard, get_skip_keyboard, get_back_keyboard
    )
    _check_subscription_func = None
    import aiogram
    if aiogram.__version__.startswith('3'):
        try:
            from aiogram.client.default import DefaultBotProperties
            bot = Bot(token=BOT_TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
        except ImportError:
            bot = Bot(token=BOT_TOKEN, parse_mode=ParseMode.HTML)
    else:
        bot = Bot(token=BOT_TOKEN, parse_mode=ParseMode.HTML)
    print(f"✅ Aiogram {aiogram.__version__}")
except ImportError as e:
    print(f"❌ Ошибка импорта aiogram: {e}")
    sys.exit(1)

# ========== PSUTIL ==========
try:
    import psutil
    PSUTIL_AVAILABLE = True
except ImportError:
    PSUTIL_AVAILABLE = False

# ========== КОНФИГУРАЦИЯ ==========
from config import get_field_label, get_field_name_only, get_visible_fields

# Старые словари для обратной совместимости (будут заменены на config)
FIELD_DB_MAP = {
    "nick": "game_nickname", "power": "power", "bm": "bm",
    "pl1": "pl1", "pl2": "pl2", "pl3": "pl3",
    "dragon": "dragon", "stands": "buffs_stands", "research": "buffs_research",
    "acceleration_buff": "acceleration_buff"  # Новое поле
}
VALID_DB_FIELDS = set(FIELD_DB_MAP.values()) | {"username"}
MAX_POWER_DRAGON, MAX_BM_PL, MAX_BUFF = 99, 999.9, 9
MAX_NICK_LENGTH, MIN_NICK_LENGTH, CACHE_TTL = 50, 2, 60
RATE_LIMIT_USER, RATE_LIMIT_ADMIN, RATE_LIMIT_WINDOW = 10, 30, 60
ACCOUNTS_PER_PAGE, MAX_BATCH_DELETE = 10, 20
cancel_restore, background_tasks_started = False, False

# ========== RATE LIMITER ==========
class RateLimiter:
    def __init__(self):
        self.requests = defaultdict(list)

    def is_limited(self, user_id: int, is_admin: bool = False) -> bool:
        now = datetime.now()
        limit = RATE_LIMIT_ADMIN if is_admin else RATE_LIMIT_USER
        window = timedelta(seconds=RATE_LIMIT_WINDOW)
        self.requests[user_id] = [t for t in self.requests[user_id] if now - t < window]
        if len(self.requests[user_id]) >= limit:
            return True
        self.requests[user_id].append(now)
        return False

rate_limiter = RateLimiter()

# ========== ДЕКОРАТОР RETRY ==========
def retry_on_db_lock(max_retries=3, delay=0.1):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            for attempt in range(max_retries):
                try:
                    return func(*args, **kwargs)
                except sqlite3.OperationalError as e:
                    if 'database is locked' in str(e) and attempt < max_retries - 1:
                        time.sleep(delay * (attempt + 1))
                        continue
                    raise
            return func(*args, **kwargs)
        return wrapper
    return decorator

# ========== БАЗА ДАННЫХ ==========
class Database:
    def __init__(self, db_name: str = DB_NAME):
        self.db_path = Path(db_name)
        self.lock = threading.RLock()
        self.cache_lock = threading.RLock()
        self.stats_cache, self.user_cache = {}, {}
        self.cache_ttl, self.last_cache_update = CACHE_TTL, 0
        self.change_counter, self.last_vacuum = 0, datetime.now()
        self.conn, self.cursor = None, None
        if self.db_path.exists():
            self._connect()

    def _connect(self):
        self.conn = sqlite3.connect(str(self.db_path), check_same_thread=False, timeout=10)
        self.conn.row_factory = sqlite3.Row
        self.cursor = self.conn.cursor()
        self._optimize()
        self._create_tables()

    def _optimize(self):
        try:
            for pragma in ["PRAGMA journal_mode=WAL", "PRAGMA synchronous=NORMAL",
                          "PRAGMA cache_size=-2000", "PRAGMA foreign_keys=ON", "PRAGMA temp_store=MEMORY"]:
                self._execute(pragma)
            self.conn.commit()
        except Exception as e:
            logger.error(f"Ошибка оптимизации БД: {e}")

    def _create_tables(self):
        self._execute('''CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER NOT NULL,
            username TEXT, game_nickname TEXT NOT NULL,
            power TEXT DEFAULT '', bm TEXT DEFAULT '',
            pl1 TEXT DEFAULT '', pl2 TEXT DEFAULT '', pl3 TEXT DEFAULT '',
            dragon TEXT DEFAULT '', buffs_stands TEXT DEFAULT '',
            buffs_research TEXT DEFAULT '', acceleration_buff TEXT DEFAULT '',
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(user_id, game_nickname))''')
        for idx in ["CREATE INDEX IF NOT EXISTS idx_user_id ON users(user_id)",
                   "CREATE INDEX IF NOT EXISTS idx_game_nickname ON users(game_nickname)",
                   "CREATE INDEX IF NOT EXISTS idx_updated_at ON users(updated_at)"]:
            try: self._execute(idx)
            except: pass
        self.conn.commit()

    def _execute(self, query: str, params: tuple = None):
        with self.lock:
            try:
                if params: self.cursor.execute(query, params)
                else: self.cursor.execute(query)
                return self.cursor
            except Exception as e:
                logger.error(f"SQL Error: {e}\nQuery: {query}")
                raise

    def _validate_field(self, field: str) -> bool:
        return field in VALID_DB_FIELDS

    def invalidate_cache(self):
        with self.cache_lock:
            self.stats_cache, self.user_cache = {}, {}
            self.last_cache_update = 0

    def get_user_accounts_cached(self, user_id: int) -> List[Dict]:
        if not self.conn: self._connect()
        cache_key = f"user_{user_id}"
        with self.cache_lock:
            if cache_key in self.user_cache:
                cache_time, cache_data = self.user_cache[cache_key]
                if time.time() - cache_time < self.cache_ttl:
                    return [dict(item) for item in cache_data] if cache_data else []
            data = self.get_user_accounts(user_id)
            self.user_cache[cache_key] = (time.time(), [dict(item) for item in data] if data else [])
            return data

    @retry_on_db_lock()
    def get_user_accounts(self, user_id: int) -> List[Dict]:
        if not self.conn: self._connect()
        try:
            self._execute("""SELECT id, game_nickname, power, bm, pl1, pl2, pl3,
                dragon, buffs_stands, buffs_research, acceleration_buff, updated_at
                FROM users WHERE user_id = ? ORDER BY updated_at DESC""", (user_id,))
            return [dict(row) for row in self.cursor.fetchall()]
        except Exception as e:
            logger.error(f"Ошибка get_user_accounts: {e}")
            return []

    @retry_on_db_lock()
    def get_account_by_id(self, account_id: int) -> Optional[Dict]:
        if not self.conn: self._connect()
        try:
            self._execute("SELECT * FROM users WHERE id = ?", (account_id,))
            row = self.cursor.fetchone()
            return dict(row) if row else None
        except Exception as e:
            logger.error(f"Ошибка get_account_by_id: {e}")
            return None

    def is_nickname_taken(self, user_id: int, nickname: str, exclude_id: int = None) -> bool:
        if not self.conn: self._connect()
        try:
            nickname = nickname.strip().lower()
            query = "SELECT id FROM users WHERE user_id = ? AND LOWER(TRIM(game_nickname)) = ?"
            params = [user_id, nickname]
            if exclude_id:
                query += " AND id != ?"
                params.append(exclude_id)
            self._execute(query, params)
            return self.cursor.fetchone() is not None
        except Exception as e:
            logger.error(f"Ошибка is_nickname_taken: {e}")
            return False

    @retry_on_db_lock()
    def create_or_update_account(self, user_id: int, username: str,
                                   game_nickname: str, field_key: str = None,
                                   value: str = None) -> Optional[Dict]:
        if not self.conn: self._connect()
        try:
            self._execute("SELECT id, game_nickname FROM users WHERE user_id = ? AND game_nickname = ?",
                         (user_id, game_nickname))
            existing = self.cursor.fetchone()
            if existing:
                account_id = existing['id']
                if field_key and value is not None:
                    db_field = FIELD_DB_MAP.get(field_key, field_key)
                    if not self._validate_field(db_field):
                        logger.error(f"Неверное поле: {db_field}")
                        return None
                    self._execute(f"""UPDATE users SET {db_field} = ?, username = ?,
                        updated_at = CURRENT_TIMESTAMP WHERE id = ?""", (value, username, account_id))
                    if field_key == "nick" and value != existing['game_nickname']:
                        self._execute("UPDATE users SET game_nickname = ? WHERE id = ?", (value, account_id))
            else:
                if field_key and value is not None:
                    db_field = FIELD_DB_MAP.get(field_key, field_key)
                    if not self._validate_field(db_field):
                        logger.error(f"Неверное поле: {db_field}")
                        return None
                    if field_key == "nick":
                        self._execute(f"""INSERT INTO users (user_id, username, game_nickname, {db_field})
                            VALUES (?, ?, ?, ?)""", (user_id, username, value, value))
                    else:
                        self._execute(f"""INSERT INTO users (user_id, username, game_nickname, {db_field})
                            VALUES (?, ?, ?, ?)""", (user_id, username, game_nickname, value))
                else:
                    self._execute("INSERT INTO users (user_id, username, game_nickname) VALUES (?, ?, ?)",
                                 (user_id, username, game_nickname))
            account_id = self.cursor.lastrowid
            self.conn.commit()
            self.invalidate_cache()
            return self.get_account_by_id(account_id)
        except sqlite3.IntegrityError:
            return None
        except Exception as e:
            logger.error(f"Ошибка create_or_update_account: {e}")
            return None

    @retry_on_db_lock()
    def delete_account(self, account_id: int) -> bool:
        if not self.conn: self._connect()
        try:
            self._execute("DELETE FROM users WHERE id = ?", (account_id,))
            self.conn.commit()
            self.invalidate_cache()
            return self.cursor.rowcount > 0
        except Exception as e:
            logger.error(f"Ошибка delete_account: {e}")
            return False

    @retry_on_db_lock()
    def get_all_accounts(self) -> List[Dict]:
        if not self.conn: self._connect()
        try:
            self._execute("""SELECT u.id, u.user_id, u.username, u.game_nickname,
                u.power, u.bm, u.pl1, u.pl2, u.pl3, u.dragon,
                u.buffs_stands, u.buffs_research, u.acceleration_buff,
                u.created_at, u.updated_at,
                p.first_name, p.last_name, p.middle_name, p.city, p.region,
                p.timezone, p.birth_day, p.birth_month, p.birth_year
                FROM users u LEFT JOIN user_profiles p ON u.user_id = p.user_id
                ORDER BY u.updated_at DESC""")
            return [dict(row) for row in self.cursor.fetchall()]
        except Exception as e:
            logger.error(f"Ошибка get_all_accounts: {e}")
            return []

    def get_stats(self) -> Dict[str, Any]:
        if not self.conn: self._connect()
        now = time.time()
        with self.cache_lock:
            if self.stats_cache and now - self.last_cache_update < self.cache_ttl:
                return self.stats_cache.copy()
        try:
            self._execute("SELECT COUNT(DISTINCT user_id) FROM users")
            unique_users = self.cursor.fetchone()[0]
            self._execute("SELECT COUNT(*) FROM users")
            total_accounts = self.cursor.fetchone()[0]
            stats = {
                "unique_users": unique_users,
                "total_accounts": total_accounts,
                "avg_accounts_per_user": round(total_accounts / unique_users, 1) if unique_users > 0 else 0
            }
            with self.cache_lock:
                self.stats_cache = stats.copy()
                self.last_cache_update = now
            return stats
        except Exception as e:
            logger.error(f"Ошибка get_stats: {e}")
            return {"unique_users": 0, "total_accounts": 0, "avg_accounts_per_user": 0}

    def update_user_last_active(self, user_id: int) -> bool:
        if not self.conn: self._connect()
        try:
            self._execute("UPDATE users SET updated_at = CURRENT_TIMESTAMP WHERE user_id = ?", (user_id,))
            self.conn.commit()
            return True
        except Exception as e:
            logger.error(f"Ошибка обновления активности: {e}")
            return False

    def check_integrity(self) -> bool:
        """Проверка целостности базы данных"""
        if not self.conn: self._connect()
        try:
            self._execute("PRAGMA integrity_check")
            result = self.cursor.fetchone()
            return result and result[0] == "ok"
        except Exception as e:
            logger.error(f"Ошибка проверки целостности БД: {e}")
            return False

    def create_backup(self, filename: str = None) -> Optional[str]:
        if not self.conn: self._connect()
        try:
            if not filename:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"backup_{timestamp}.db"
            filepath = BACKUP_DIR / filename
            with self.lock:
                self.conn.commit()
                self.cursor.execute("PRAGMA integrity_check")
                if self.cursor.fetchone()[0] != "ok":
                    self.cursor.execute("REINDEX")
                    self.conn.commit()
                backup_conn = sqlite3.connect(str(filepath))
                self.conn.backup(backup_conn)
                backup_conn.close()
                logger.info(f"✅ Бэкап создан: {filepath}")
                return str(filepath)
        except Exception as e:
            logger.error(f"❌ Ошибка создания бэкапа: {e}")
            return None

    def export_to_csv(self, filename: str = None) -> Optional[str]:
        if not self.conn: self._connect()
        try:
            if not filename:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"export_{timestamp}.csv"
            filepath = EXPORT_DIR / filename
            accounts = self.get_all_accounts()
            if not accounts: return None

            # Формируем заголовки и данные динамически на основе видимых полей
            visible = get_visible_fields()
            headers = ["№", "Ник"]
            for key in visible:
                name = get_field_label(key)
                if name:
                    headers.append(name)

            with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f, delimiter=';')
                writer.writerow(headers)
                for i, acc in enumerate(accounts, 1):
                    row = [i, acc.get('game_nickname', '')]
                    for key in visible:
                        db_key = FIELD_DB_MAP.get(key, key)
                        row.append(acc.get(db_key, ''))
                    writer.writerow(row)
            logger.info(f"✅ Экспорт CSV: {filepath}")
            return str(filepath)
        except Exception as e:
            logger.error(f"❌ Ошибка экспорта CSV: {e}")
            return None

    def export_to_excel(self, filename: str = None) -> Optional[str]:
        if not self.conn: self._connect()
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            logger.error("❌ openpyxl не установлен")
            return None
        try:
            if not filename:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"export_{timestamp}.xlsx"
            filepath = EXPORT_DIR / filename
            accounts = self.get_all_accounts()
            if not accounts: return None
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Игроки"
            headers = ["№", "Группа", "Ник", "ФИО", "Город", "Дата рождения", "Часовой пояс",
                      "Эл", "БМ", "Пл1", "Пл2", "Пл3", "Др", "БС", "БИ", "ID имя", "ID номер", "Время", "Дата"]
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font_white = Font(bold=True, color="FFFFFF")
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            def format_number(val):
                if not val or val == '—': return ''
                try: return round(float(val.replace(',', '.')) * 10) / 10
                except: return val
            for i, acc in enumerate(accounts, 1):
                row_data = [i, '', acc.get('game_nickname', ''),
                    f"{acc.get('last_name','')} {acc.get('first_name','')}".strip(),
                    acc.get('city', ''), '', acc.get('timezone', 'Europe/Moscow'),
                    acc.get('power', ''), format_number(acc.get('bm', '')),
                    format_number(acc.get('pl1', '')), format_number(acc.get('pl2', '')),
                    format_number(acc.get('pl3', '')), acc.get('dragon', ''),
                    acc.get('buffs_stands', ''), acc.get('buffs_research', ''),
                    f"@{acc.get('username','')}" if acc.get('username') else '',
                    acc.get('user_id', ''), '', '']
                for col, value in enumerate(row_data, 1):
                    ws.cell(row=i+1, column=col, value=value)
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 15
            wb.save(filepath)
            logger.info(f"✅ Экспорт Excel: {filepath}")
            return str(filepath)
        except Exception as e:
            logger.error(f"❌ Ошибка экспорта в Excel: {e}")
            return None

    def close(self):
        try:
            with self.lock:
                if self.conn:
                    self.conn.commit()
                    self.conn.close()
                    self.conn = None
        except: pass

db = Database()
if not db.conn:
    db._connect()
    print("✅ Соединение с БД создано")

# ========== ФУНКЦИЯ ПРОВЕРКИ ПОДПИСКИ ==========
async def check_subscription(user_id: int) -> bool:
    global _check_subscription_func
    _check_subscription_func = check_subscription
    if not TARGET_CHAT_ID: return True
    try:
        member = await bot.get_chat_member(chat_id=TARGET_CHAT_ID, user_id=user_id)
        return member.status in ['creator', 'administrator', 'member']
    except: return False

# ========== ИНИЦИАЛИЗАЦИЯ МОДУЛЕЙ ==========
profile_db = ProfileDB(db)
city_db = CityDatabase()
import handlers.profile
handlers.profile.profile_db = profile_db
handlers.profile.db = db
handlers.profile._check_subscription_func = check_subscription
print("✅ Экземпляры переданы в handlers.profile")

# ========== FSM ==========
storage = MemoryStorage()
dp = Dispatcher(storage=storage)
router = Router()
dp.include_router(profile.router)
dp.include_router(router)

class EditState(StatesGroup):
    waiting_field_value = State()
    step_by_step = State()
    waiting_search_query = State()
    waiting_batch_delete = State()
    waiting_for_backup = State()
    batch_selection = State()

# ========== КЛАВИАТУРЫ ==========
def is_admin(user_id: int) -> bool:
    return user_id in ADMIN_IDS

def get_main_kb(user_id: int) -> ReplyKeyboardMarkup:
    kb = [[KeyboardButton(text="📊 Мои аккаунты"), KeyboardButton(text="📤 Отправить в группу")],
          [KeyboardButton(text="👤 Мой профиль")]]
    if is_admin(user_id):
        kb.append([KeyboardButton(text="👑 Админ-панель")])
    return ReplyKeyboardMarkup(keyboard=kb, resize_keyboard=True)

def get_numeric_kb(decimal: bool = True) -> ReplyKeyboardMarkup:
    if decimal:
        kb = [[KeyboardButton(text=str(i)) for i in range(1,4)],
              [KeyboardButton(text=str(i)) for i in range(4,7)],
              [KeyboardButton(text=str(i)) for i in range(7,10)],
              [KeyboardButton(text="0"), KeyboardButton(text=","), KeyboardButton(text="⌫")],
              [KeyboardButton(text="🏁 Завершить"), KeyboardButton(text="⏭ Пропустить"), KeyboardButton(text="✅ Готово")]]
    else:
        kb = [[KeyboardButton(text=str(i)) for i in range(1,4)],
              [KeyboardButton(text=str(i)) for i in range(4,7)],
              [KeyboardButton(text=str(i)) for i in range(7,10)],
              [KeyboardButton(text="0"), KeyboardButton(text="⌫"), KeyboardButton(text="⌫")],
              [KeyboardButton(text="🏁 Завершить"), KeyboardButton(text="⏭ Пропустить"), KeyboardButton(text="✅ Готово")]]
    return ReplyKeyboardMarkup(keyboard=kb, resize_keyboard=True)

def get_cancel_kb() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(keyboard=[[KeyboardButton(text="🚫 Отмена")]], resize_keyboard=True)

def get_accounts_kb(accounts: List[Dict]) -> InlineKeyboardMarkup:
    buttons = [[InlineKeyboardButton(text=f"👤 {acc.get('game_nickname','')[:20]}",
                                     callback_data=f"select_{acc.get('id')}")]
               for acc in accounts[:10] if acc.get('id')]
    buttons.append([InlineKeyboardButton(text="➕ Новый аккаунт", callback_data="new_account")])
    buttons.append([InlineKeyboardButton(text="⬅️ Назад", callback_data="menu")])
    return InlineKeyboardMarkup(inline_keyboard=buttons)

def get_account_actions_kb(account_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✏️ Изменить ник", callback_data=f"edit_nick_{account_id}")],
        [InlineKeyboardButton(text="📝 Редактировать", callback_data=f"edit_{account_id}")],
        [InlineKeyboardButton(text="🔄 Пошагово", callback_data=f"step_{account_id}")],
        [InlineKeyboardButton(text="📤 Отправить", callback_data=f"send_{account_id}")],
        [InlineKeyboardButton(text="🗑️ Удалить", callback_data=f"delete_{account_id}")],
        [InlineKeyboardButton(text="⬅️ Назад", callback_data="my_accounts")]
    ])

def get_send_kb(accounts: List[Dict]) -> InlineKeyboardMarkup:
    buttons = [[InlineKeyboardButton(text=f"📤 {acc.get('game_nickname','')[:20]}",
                                     callback_data=f"send_{acc.get('id')}")]
               for acc in accounts[:10] if acc.get('id')]
    buttons.append([InlineKeyboardButton(text="⬅️ Отмена", callback_data="menu")])
    return InlineKeyboardMarkup(inline_keyboard=buttons)

def get_admin_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📋 Таблица", callback_data="admin_table_1")],
        [InlineKeyboardButton(text="📤 Экспорт CSV", callback_data="admin_export")],
        [InlineKeyboardButton(text="📊 Экспорт Excel", callback_data="admin_export_excel")],
        [InlineKeyboardButton(text="🗄️ Управление БД", callback_data="db_management")],
        [InlineKeyboardButton(text="🔍 Поиск", callback_data="admin_search")],
        [InlineKeyboardButton(text="🗑️ Пакетное удаление", callback_data="admin_batch")],
        [InlineKeyboardButton(text="📊 Статистика", callback_data="admin_stats")],
        [InlineKeyboardButton(text="🔄 Обновить", callback_data="admin_refresh")],
        [InlineKeyboardButton(text="🏠 Меню", callback_data="menu")]
    ])

def get_db_management_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="💾 Сохранить бэкап", callback_data="db_backup")],
        [InlineKeyboardButton(text="📥 Восстановить из бэкапа", callback_data="db_restore_menu")],
        [InlineKeyboardButton(text="📤 Загрузить с ПК", callback_data="db_restore_pc")],
        [InlineKeyboardButton(text="🧹 Очистка (14 дней)", callback_data="admin_cleanup")],
        [InlineKeyboardButton(text="⬅️ Назад", callback_data="admin_back")]
    ])

def get_edit_fields_kb(account_id: int) -> InlineKeyboardMarkup:
    buttons = []
    for key in get_visible_fields():
        if key == "nick":
            continue
        name = get_field_label(key)
        if name:
            buttons.append([InlineKeyboardButton(text=name, callback_data=f"field_{account_id}_{key}")])
    buttons.append([InlineKeyboardButton(text="⬅️ Назад", callback_data=f"select_{account_id}")])
    return InlineKeyboardMarkup(inline_keyboard=buttons)

# ========== ФОРМАТТЕРЫ ==========
def format_power(value: str) -> str:
    if not value or value == '—': return ' —'
    try: return f"{min(int(value.replace(',','').strip()), MAX_POWER_DRAGON):2d}"
    except: return ' —'

def format_bm(value: str) -> str:
    if not value or value == '—': return '   —'
    try: return f"{round(min(float(value.replace(',', '.')), MAX_BM_PL), 1):5.1f}".replace('.', ',')
    except: return '   —'

def format_pl(value: str) -> str: return format_bm(value)

def format_dragon(value: str) -> str: return format_power(value)

def format_buff(value: str) -> str:
    if not value or value == '—': return '—'
    try: return str(min(int(value.replace(',', '').strip()), MAX_BUFF))
    except: return '—'

def format_timezone_offset(tzid: str) -> str:
    offsets = {'Europe/Kaliningrad': 'MSK-1', 'Europe/Moscow': 'MSK', 'Europe/Volgograd': 'MSK',
               'Europe/Kirov': 'MSK', 'Europe/Astrakhan': 'MSK+1', 'Europe/Samara': 'MSK+1',
               'Europe/Saratov': 'MSK+1', 'Europe/Ulyanovsk': 'MSK+1', 'Asia/Yekaterinburg': 'MSK+2',
               'Asia/Omsk': 'MSK+3', 'Asia/Novosibirsk': 'MSK+4', 'Asia/Barnaul': 'MSK+4',
               'Asia/Tomsk': 'MSK+4', 'Asia/Novokuznetsk': 'MSK+4', 'Asia/Krasnoyarsk': 'MSK+4',
               'Asia/Irkutsk': 'MSK+5', 'Asia/Chita': 'MSK+6', 'Asia/Yakutsk': 'MSK+6',
               'Asia/Khandyga': 'MSK+6', 'Asia/Vladivostok': 'MSK+7', 'Asia/Ust-Nera': 'MSK+7',
               'Asia/Magadan': 'MSK+8', 'Asia/Sakhalin': 'MSK+8', 'Asia/Srednekolymsk': 'MSK+8',
               'Asia/Kamchatka': 'MSK+9', 'Asia/Anadyr': 'MSK+9'}
    return offsets.get(tzid, tzid.replace('Europe/', '').replace('Asia/', ''))

def format_accounts_table(accounts: List[Dict], start: int = 0) -> str:
    text = "<code>\n"
    for i, acc in enumerate(accounts, start + 1):
        nick = html.escape((acc.get('game_nickname') or '—')[:20])
        text += f"{i:2d}. {nick}\n"

        # Динамическое формирование строки на основе видимых полей
        line_parts = []
        visible = get_visible_fields()

        for field_key in visible:
            label = get_field_label(field_key)
            if not label:
                continue

            value = acc.get(field_key, '—')
            if value is None or value == '':
                value = '—'

            # Форматирование значения в зависимости от типа поля
            if field_key == 'power':
                formatted_val = format_power(value)
            elif field_key == 'bm':
                formatted_val = format_bm(value)
            elif field_key in ['pl1', 'pl2', 'pl3']:
                formatted_val = format_pl(value)
            else:
                formatted_val = format_buff(value) if 'buff' in field_key or field_key == 'dragon' else value

            line_parts.append(f"{label} {formatted_val}")

        text += " ".join(line_parts) + "\n"
    text += "</code>"
    return text

def format_account_data(acc: Dict) -> str:
    if not acc: return "❌ Аккаунт не найден"
    nick = acc.get('game_nickname', 'Без имени')
    text = f"<b>📋 Аккаунт: {html.escape(nick)}</b>\n"
    for key in get_visible_fields():
        label = get_field_label(key)
        if not label:
            continue
        # Получаем техническое имя поля из БД (если нужно маппирование)
        db_key = FIELD_DB_MAP.get(key, key)
        val = acc.get(db_key, '')
        text += f"<b>{label}:</b> {html.escape(str(val)) if val else '—'}\n"
    return text

# ========== ВАЛИДАЦИЯ ==========
def validate_numeric_input(field: str, value: str) -> tuple:
    try:
        if field in ["bm", "pl1", "pl2", "pl3"]:
            parts = value.split(',')
            if len(parts) > 2: return False, "❌ Неверный формат", value
            num = float(value.replace(',', '.'))
            if num > MAX_BM_PL: return False, f"❌ Максимум: {MAX_BM_PL}", value
        elif field in ["power", "dragon"]:
            cleaned = value.replace(',', '')
            if not cleaned.isdigit(): return False, "❌ Введите целое число", value
            num = int(cleaned)
            if num > MAX_POWER_DRAGON: return False, f"❌ Максимум: {MAX_POWER_DRAGON}", value
            value = cleaned
        elif field in ["stands", "research"]:
            cleaned = value.replace(',', '')
            if not cleaned.isdigit(): return False, "❌ Введите число 0-9", value
            num = int(cleaned)
            if num > MAX_BUFF: return False, f"❌ Максимум: {MAX_BUFF}", value
            value = cleaned
        return True, "", value
    except ValueError:
        return False, "❌ Введите корректное число", value

# ========== SAFE SEND ==========
async def safe_send(obj, text: str, **kwargs):
    MAX_LEN = 4096
    try:
        if len(text) <= MAX_LEN:
            if isinstance(obj, CallbackQuery) and obj.message:
                try: await obj.message.edit_text(text, **kwargs)
                except: await obj.message.answer(text, **kwargs)
            elif isinstance(obj, Message): await obj.answer(text, **kwargs)
        else:
            parts = [text[i:i+MAX_LEN] for i in range(0, len(text), MAX_LEN)]
            for i, part in enumerate(parts):
                if i == 0 and isinstance(obj, CallbackQuery) and obj.message:
                    try: await obj.message.edit_text(part, **kwargs)
                    except: await obj.message.answer(part, **kwargs)
                elif isinstance(obj, Message): await obj.answer(part, **kwargs)
    except Exception as e: logger.error(f"Safe send error: {e}")

# ========== КОМАНДЫ ==========
@router.message(Command("start"))
async def start_cmd(message: Message):
    user_id = message.from_user.id
    if rate_limiter.is_limited(user_id, is_admin(user_id)):
        await message.answer("⏳ Слишком много запросов"); return
    if profile_db:
        profile_db.update_last_active(user_id)
        db.update_user_last_active(user_id)
    accounts = db.get_user_accounts_cached(user_id)
    text = "🎮 <b>Бот для сбора игровых данных</b>\n👋 Добро пожаловать!"
    if not accounts:
        text += "\nУ вас нет аккаунтов. Нажмите «📊 Мои аккаунты»"
    else:
        text += f"\n📊 Аккаунтов: {len(accounts)}"
    await message.answer(text, reply_markup=get_main_kb(user_id))

@router.message(Command("help"))
async def help_cmd(message: Message):
    await message.answer("📖 <b>Помощь</b>\n/start - Запуск\n/help - Помощь\n/cancel - Отмена")

@router.message(Command("cancel"))
async def cancel_cmd(message: Message, state: FSMContext):
    global cancel_restore
    if is_admin(message.from_user.id): cancel_restore = True
    await state.clear()
    await message.answer("❌ Отменено", reply_markup=get_main_kb(message.from_user.id))

@router.message(Command("myid"))
async def myid_cmd(message: Message):
    await message.answer(f"🆔 <b>Ваш ID:</b> <code>{message.from_user.id}</code>")

@router.message(Command("admin"))
async def admin_cmd(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("🚫 Только для админов"); return
    stats = db.get_stats()
    await message.answer(f"👑 <b>Админ-панель</b>\n👥 {stats['unique_users']}\n🎮 {stats['total_accounts']}",
                        reply_markup=get_admin_kb())

# ========== ОСНОВНЫЕ КНОПКИ ==========
@router.message(F.text == "📊 Мои аккаунты")
async def my_accounts(message: Message):
    user_id = message.from_user.id
    accounts = db.get_user_accounts(user_id)
    if not accounts:
        await message.answer("📋 У вас нет аккаунтов",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="➕ Создать", callback_data="new_account")],
                [InlineKeyboardButton(text="🏠 Меню", callback_data="menu")]]))
        return
    text = "<b>📋 Ваши аккаунты:</b>\n" + format_accounts_table(accounts)
    await safe_send(message, text, reply_markup=get_accounts_kb(accounts))

@router.message(F.text == "👤 Мой профиль")
async def my_profile_button(message: Message, state: FSMContext):
    from handlers.profile import cmd_profile
    await cmd_profile(message)

@router.message(F.text == "📤 Отправить в группу")
async def send_menu(message: Message):
    if not TARGET_CHAT_ID:
        await message.answer("❌ Отправка не настроена"); return
    accounts = db.get_user_accounts_cached(message.from_user.id)
    if not accounts:
        await message.answer("❌ Сначала создайте аккаунт"); return
    await message.answer("📤 Выберите аккаунт:", reply_markup=get_send_kb(accounts))

@router.message(F.text == "👑 Админ-панель")
async def admin_panel_msg(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("🚫 Доступ запрещен"); return
    stats = db.get_stats()
    await message.answer(f"👑 <b>Админ-панель</b>\n👥 {stats['unique_users']}", reply_markup=get_admin_kb())

# ========== ПОШАГОВОЕ ЗАПОЛНЕНИЕ ==========
@router.callback_query(F.data.startswith("step_"))
async def step_start(callback: CallbackQuery, state: FSMContext):
    account_id = int(callback.data.split("_")[1])
    account = db.get_account_by_id(account_id)
    if not account:
        await callback.answer("❌ Аккаунт не найден", show_alert=True); return
    steps = [k for k in get_visible_fields() if k != "nick"]
    await callback.message.edit_text(f"🔄 <b>ПОШАГОВОЕ ЗАПОЛНЕНИЕ</b>\n👤 {account['game_nickname']}\n📊 {len(steps)} полей")
    await state.update_data(step_account=account_id, step_index=0, step_steps=steps, step_data={}, step_temp="")
    await step_next(callback.message, state)
    await callback.answer()

async def step_next(msg_or_cb, state: FSMContext):
    data = await state.get_data()
    account_id = data.get("step_account")
    idx = data.get("step_index", 0)
    steps = data.get("step_steps", [])
    if idx >= len(steps):
        await step_finish(msg_or_cb, state); return
    field = steps[idx]
    account = db.get_account_by_id(account_id)
    if not account: await state.clear(); return
    name = get_field_label(field) or field
    current = account.get(FIELD_DB_MAP.get(field, field), '')
    text = f"🔄 <b>ШАГ {idx+1} ИЗ {len(steps)}</b>\n👤 {account['game_nickname']}\n📌 {name}\n💾 Текущее: {current or '—'}"
    if isinstance(msg_or_cb, Message): await msg_or_cb.answer(text)
    else: await msg_or_cb.message.edit_text(text)
    if field in ["bm", "pl1", "pl2", "pl3"]:
        kb, prompt = get_numeric_kb(decimal=True), f"📝 Введите число для «{name}» (можно с запятой):"
    elif field in ["power", "dragon", "stands", "research", "acceleration_buff"]:
        kb, prompt = get_numeric_kb(decimal=False), f"📝 Введите целое число для «{name}»:"
    else:
        kb, prompt = get_cancel_kb(), f"📝 Введите значение:"
    await msg_or_cb.answer(prompt, reply_markup=kb)
    await state.set_state(EditState.step_by_step)
    await state.update_data(step_field=field, step_temp="")

@router.message(EditState.step_by_step)
async def step_input(message: Message, state: FSMContext):
    data = await state.get_data()
    field = data.get("step_field")
    step_data = data.get("step_data", {})
    step_temp = data.get("step_temp", "")
    field_name = get_field_label(field) or field

    if message.text == "🚫 Отмена":
        await message.answer("❌ Отменено", reply_markup=get_main_kb(message.from_user.id))
        await state.clear(); return
    if message.text == "🏁 Завершить":
        await step_finish(message, state, early=True); return
    if message.text == "⏭ Пропустить":
        await message.answer(f"⏭ Поле «{field_name}» пропущено")
        await state.update_data(step_index=data.get("step_index", 0) + 1, step_temp="")
        await step_next(message, state); return

    # ===== ПРОВЕРКА: КНОПКА ИЛИ КЛАВИАТУРА? =====
    is_single_char = len(message.text) == 1 and message.text in ["0","1","2","3","4","5","6","7","8","9",",","⌫"]

    if is_single_char:
        # ===== РЕЖИМ КНОПОК =====
        if message.text in ["0","1","2","3","4","5","6","7","8","9"]:
            step_temp += message.text
            await state.update_data(step_temp=step_temp)
            await message.answer(f"📝 Текущее: {step_temp}"); return
        if message.text == ",":
            if field in ["bm","pl1","pl2","pl3"] and "," not in step_temp:
                step_temp += ","
                await state.update_data(step_temp=step_temp)
                await message.answer(f"📝 Текущее: {step_temp}")
            else:
                await message.answer("📝 Введите целое число без запятой")
            return
        if message.text == "⌫":
            step_temp = step_temp[:-1] if step_temp else ""
            await state.update_data(step_temp=step_temp)
            await message.answer(f"📝 Текущее: {step_temp}" if step_temp else "📝 Очищено"); return
        if message.text == "✅ Готово":
            if step_temp:
                value = step_temp
                await state.update_data(step_temp="")
            else:
                await message.answer("❌ Нет значения. Используйте кнопки."); return
        else:
            await message.answer("❌ Используйте кнопки или нажмите ✅ Готово"); return
    else:
        # ===== РУЧНОЙ ВВОД С КЛАВИАТУРЫ =====
        value = message.text.strip()
        await state.update_data(step_temp="")
        print(f"⌨️ Ручной ввод: '{value}'")
        if not value:
            await message.answer("❌ Значение не может быть пустым"); return

    # ===== ВАЛИДАЦИЯ =====
    if field in ["power","bm","dragon","stands","research","pl1","pl2","pl3"]:
        value = value.replace('.', ',')
        success, error_msg, cleaned_value = validate_numeric_input(field, value)
        if not success:
            kb = get_numeric_kb(decimal=True) if field in ["bm","pl1","pl2","pl3"] else get_numeric_kb(decimal=False)
            await message.answer(error_msg, reply_markup=kb); return
        value = cleaned_value

    # ===== СОХРАНЕНИЕ И ПЕРЕХОД =====
    step_data[field] = value
    await message.answer(f"✅ {field_name}: {value}")
    await state.update_data(step_data=step_data, step_index=data.get("step_index", 0) + 1, step_temp="")
    await step_next(message, state)

async def step_finish(msg_or_cb, state: FSMContext, early=False):
    data = await state.get_data()
    account_id = data.get("step_account")
    step_data = data.get("step_data", {})
    account = db.get_account_by_id(account_id)
    if not account: await state.clear(); return
    user_id = msg_or_cb.from_user.id
    username = msg_or_cb.from_user.username or f"user_{user_id}"
    updated = []
    for field, value in step_data.items():
        if value and value.strip():
            db.create_or_update_account(user_id, username, account['game_nickname'], field, value)
            updated.append(get_field_label(field) or field)
    text = "🏁 <b>ПРЕРВАНО</b>" if early else "✅ <b>ЗАВЕРШЕНО!</b>"
    text += f"\n👤 {account['game_nickname']}\n📊 Обновлено: {len(updated)}"
    if isinstance(msg_or_cb, Message):
        await msg_or_cb.answer(text, reply_markup=get_main_kb(user_id))
    else:
        await msg_or_cb.message.edit_text(text,
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="🏠 Меню", callback_data="menu")]]))
    await state.clear()

# ========== ОБРАБОТКА ВВОДА ==========
@router.message(EditState.waiting_field_value)
async def process_input(message: Message, state: FSMContext):
    user_id = message.from_user.id
    username = message.from_user.username or f"user_{user_id}"
    data = await state.get_data()
    field = data.get("field")
    new = data.get("new", False)
    account_id = data.get("account_id")

    if message.text in ["🚫 Отмена", "🏁 Завершить", "⏭ Пропустить"]:
        if message.text == "🚫 Отмена" or message.text == "🏁 Завершить":
            await message.answer("❌ Отменено" if message.text == "🚫 Отмена" else "🏁 Завершено",
                               reply_markup=get_main_kb(user_id))
        else:
            await message.answer(f"⏭ Поле «{get_field_label(field) or field}» пропущено",
                               reply_markup=get_main_kb(user_id))
        await state.clear(); return

    value = message.text.strip()
    field_name = get_field_label(field) or field

    if field == "nick":
        if not value or len(value) < MIN_NICK_LENGTH or len(value) > MAX_NICK_LENGTH:
            await message.answer(f"❌ Ник должен быть от {MIN_NICK_LENGTH} до {MAX_NICK_LENGTH} символов",
                               reply_markup=get_cancel_kb()); return
        if db.is_nickname_taken(user_id, value, account_id):
            await message.answer(f"❌ Ник '{value}' уже используется", reply_markup=get_cancel_kb()); return
        if new:
            acc = db.create_or_update_account(user_id, username, value)
            if acc:
                await message.answer(f"✅ Аккаунт создан: {value}", reply_markup=get_main_kb(user_id))
                await state.clear(); return
        elif account_id:
            acc = db.get_account_by_id(account_id)
            if acc:
                db.create_or_update_account(user_id, username, acc['game_nickname'], "nick", value)
                await message.answer(f"✅ Ник изменен: {value}", reply_markup=get_main_kb(user_id))
                await state.clear(); return

    if field in ["power","bm","dragon","stands","research","pl1","pl2","pl3"]:
        value = value.replace('.', ',')
        success, error_msg, cleaned_value = validate_numeric_input(field, value)
        if not success:
            kb = get_numeric_kb(decimal=True) if field in ["bm","pl1","pl2","pl3"] else get_numeric_kb(decimal=False)
            await message.answer(error_msg, reply_markup=kb); return
        value = cleaned_value
        if account_id:
            account = db.get_account_by_id(account_id)
            if account:
                db.create_or_update_account(user_id, username, account['game_nickname'], field, value)
                await message.answer(f"✅ {field_name}: {value}", reply_markup=get_main_kb(user_id))
                await state.clear(); return

# ========== ОБРАБОТКА ФАЙЛОВ ==========
@router.message(EditState.waiting_for_backup, F.document)
async def handle_backup_file(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id): await state.clear(); return
    if not message.document.file_name.endswith('.db'):
        await message.answer("❌ Нужен файл с расширением .db"); await state.clear(); return
    status_msg = await message.answer("🔄 Загружаю и восстанавливаю бэкап...")
    try:
        file = await bot.get_file(message.document.file_id)
        downloaded_file = await bot.download_file(file.file_path)
        temp_path = BACKUP_DIR / f"restore_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
        with open(temp_path, 'wb') as f: f.write(downloaded_file.getvalue())
        current_backup = BACKUP_DIR / f"before_restore_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
        if db.db_path.exists(): shutil.copy2(db.db_path, current_backup)
        db.close()
        if db.db_path.exists(): os.remove(db.db_path)
        shutil.copy2(temp_path, db.db_path)
        db._connect()
        if db.check_integrity():
            db._execute("SELECT COUNT(*) FROM users")
            users_count = db.cursor.fetchone()[0]
            await status_msg.edit_text(f"✅ База данных восстановлена!\n📊 Аккаунтов: {users_count}")
        else:
            if current_backup.exists():
                db.close()
                shutil.copy2(current_backup, db.db_path)
                db._connect()
            await status_msg.edit_text("❌ Файл поврежден. Восстановлена предыдущая БД.")
    except Exception as e:
        logger.error(f"Ошибка восстановления: {e}")
        await status_msg.edit_text(f"❌ Ошибка: {e}")
        try: db._connect()
        except: pass
    finally:
        try:
            if 'temp_path' in locals() and temp_path.exists(): temp_path.unlink()
        except: pass
        await state.clear()

# ========== ОБЩИЙ ХЕНДЛЕР ==========
@router.message(F.chat.type == "private")
async def any_message(message: Message, state: FSMContext):
    current_state = await state.get_state()
    if current_state == EditState.waiting_search_query:
        await process_search(message, state); return
    if current_state is not None: return
    if message.text in ["📊 Мои аккаунты", "📤 Отправить в группу", "👑 Админ-панель"]: return
    user_id = message.from_user.id
    if rate_limiter.is_limited(user_id, is_admin(user_id)):
        await message.answer("⏳ Слишком много запросов"); return
    accounts = db.get_user_accounts_cached(user_id)
    if accounts:
        await message.answer("🏠 <b>Главное меню</b>", reply_markup=get_main_kb(user_id)); return
    if message.text != "/start":
        await message.answer("👋 <b>Привет!</b> Введи /start",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="🚀 Запустить", callback_data="force_start")]]))

# ========== НАВИГАЦИЯ ==========
@router.callback_query(F.data == "force_start")
async def force_start(callback: CallbackQuery):
    await callback.answer()
    await start_cmd(callback.message)

@router.callback_query(F.data == "menu")
async def menu_cb(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.edit_text("🏠 Главное меню", reply_markup=None)
    await callback.message.answer("🏠 Главное меню", reply_markup=get_main_kb(callback.from_user.id))
    await callback.answer()

@router.callback_query(F.data == "new_account")
async def new_account(callback: CallbackQuery, state: FSMContext):
    user_id = callback.from_user.id
    if not await check_subscription(user_id):
        await callback.message.edit_text("❌ Нужно быть подписчиком группы", reply_markup=get_main_kb(user_id))
        await callback.answer(); return
    await callback.message.edit_text("➕ <b>Создание аккаунта</b>\nВведите игровой ник:")
    await callback.message.answer(f"📝 Введите ник ({MIN_NICK_LENGTH}-{MAX_NICK_LENGTH}):", reply_markup=get_cancel_kb())
    await state.set_state(EditState.waiting_field_value)
    await state.update_data(field="nick", new=True, temp="")
    await callback.answer()

@router.callback_query(F.data.startswith("select_"))
async def select_account(callback: CallbackQuery):
    try: account_id = int(callback.data.split("_")[1])
    except: await callback.answer("❌ Неверный ID", show_alert=True); return
    account = db.get_account_by_id(account_id)
    if not account: await callback.answer("❌ Аккаунт не найден", show_alert=True); return
    await callback.message.edit_text(format_account_data(account), reply_markup=get_account_actions_kb(account_id))
    await callback.answer()

@router.callback_query(F.data.startswith("edit_"))
async def edit_account(callback: CallbackQuery):
    try: account_id = int(callback.data.split("_")[1])
    except: await callback.answer("❌ Неверный ID", show_alert=True); return
    account = db.get_account_by_id(account_id)
    if not account: await callback.answer("❌ Аккаунт не найден", show_alert=True); return
    await callback.message.edit_text(f"✏️ <b>Редактирование</b> {account['game_nickname']}",
                                    reply_markup=get_edit_fields_kb(account_id))
    await callback.answer()

@router.callback_query(F.data.startswith("field_"))
async def edit_field(callback: CallbackQuery, state: FSMContext):
    parts = callback.data.split("_")
    if len(parts) < 3: await callback.answer("❌ Неверный формат", show_alert=True); return
    try: account_id = int(parts[1])
    except: await callback.answer("❌ Неверный ID", show_alert=True); return
    field = parts[2]
    if field not in FIELDS: await callback.answer("❌ Неверное поле", show_alert=True); return
    account = db.get_account_by_id(account_id)
    if not account: await callback.answer("❌ Аккаунт не найден", show_alert=True); return
    db_field = FIELD_DB_MAP.get(field, field)
    current = account.get(db_field, '')
    name = get_field_label(field) or field
    await callback.message.edit_text(f"✏️ <b>{name}</b>\nТекущее: {current or '—'}\nВведите новое значение:")
    if field in ["bm","pl1","pl2","pl3"]:
        await callback.message.answer("📝 Введите число (можно с запятой):", reply_markup=get_numeric_kb(decimal=True))
    elif field in ["power","dragon"]:
        await callback.message.answer(f"📝 Введите целое число (0-{MAX_POWER_DRAGON}):", reply_markup=get_numeric_kb(decimal=False))
    elif field in ["stands","research"]:
        await callback.message.answer(f"📝 Введите число (0-{MAX_BUFF}):", reply_markup=get_numeric_kb(decimal=False))
    else:
        await callback.message.answer("📝 Введите значение:", reply_markup=get_cancel_kb())
    await state.set_state(EditState.waiting_field_value)
    await state.update_data(field=field, account_id=account_id, temp="")
    await callback.answer()

@router.callback_query(F.data.startswith("delete_"))
async def delete_account(callback: CallbackQuery):
    try: account_id = int(callback.data.split("_")[1])
    except: await callback.answer("❌ Неверный ID", show_alert=True); return
    account = db.get_account_by_id(account_id)
    if not account: await callback.answer("❌ Аккаунт не найден", show_alert=True); return
    await callback.message.edit_text(f"🗑️ Удалить {account['game_nickname']}?",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✅ Да", callback_data=f"confirm_delete_{account_id}")],
            [InlineKeyboardButton(text="❌ Нет", callback_data=f"select_{account_id}")]]))
    await callback.answer()

@router.callback_query(F.data.startswith("confirm_delete_"))
async def confirm_delete(callback: CallbackQuery):
    try: account_id = int(callback.data.split("_")[2])
    except: await callback.answer("❌ Неверный ID", show_alert=True); return
    account = db.get_account_by_id(account_id)
    if not account: await callback.answer("❌ Аккаунт не найден", show_alert=True); return
    if db.delete_account(account_id):
        db.invalidate_cache()
        await callback.message.edit_text(f"✅ Удален: {account['game_nickname']}",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="🏠 Меню", callback_data="menu")]]))
    else:
        await callback.message.edit_text("❌ Ошибка удаления",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="⬅️ Назад", callback_data=f"select_{account_id}")]]))
    await callback.answer()

@router.callback_query(F.data.startswith("send_"))
async def send_account(callback: CallbackQuery):
    if not TARGET_CHAT_ID: await callback.answer("❌ Отправка не настроена", show_alert=True); return
    try: account_id = int(callback.data.split("_")[1])
    except: await callback.answer("❌ Неверный ID", show_alert=True); return
    account = db.get_account_by_id(account_id)
    if not account: await callback.answer("❌ Аккаунт не найден", show_alert=True); return
    text = f"📊 <b>Данные:</b> {account['game_nickname']}\n"
    for key in get_visible_fields():
        name = get_field_label(key)
        if key == "nick": continue
        val = account.get(FIELD_DB_MAP.get(key, key), '')
        if val and val != '—': text += f"<b>{name}:</b> {val}\n"
    text += f"\n👤 От: @{callback.from_user.username or 'пользователь'}"
    try:
        if USE_TOPIC and TARGET_TOPIC_ID:
            await bot.send_message(chat_id=TARGET_CHAT_ID, message_thread_id=TARGET_TOPIC_ID, text=text)
        else:
            await bot.send_message(chat_id=TARGET_CHAT_ID, text=text)
        await callback.message.edit_text(f"✅ Отправлено: {account['game_nickname']}",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="🏠 Меню", callback_data="menu")]]))
        await callback.answer("✅ Отправлено!")
    except Exception as e:
        logger.error(f"Send error: {e}")
        await callback.answer("❌ Ошибка отправки", show_alert=True)

# ========== УПРАВЛЕНИЕ БД ==========
@router.callback_query(F.data == "db_management")
async def db_management_menu(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("🚫 Доступ запрещен", show_alert=True); return
    stats = db.get_stats()
    await callback.message.edit_text(f"🗄️ <b>Управление БД</b>\n👥 {stats['unique_users']}\n🎮 {stats['total_accounts']}",
                                    reply_markup=get_db_management_kb())
    await callback.answer()

@router.callback_query(F.data == "db_backup")
async def db_backup_handler(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("🚫 Доступ запрещен", show_alert=True); return
    await callback.message.edit_text("🔄 Создание бэкапа...")
    path = await asyncio.to_thread(db.create_backup)
    if path and Path(path).exists():
        try:
            await bot.send_document(chat_id=callback.from_user.id, document=FSInputFile(path),
                                   caption=f"💾 Бэкап от {datetime.now().strftime('%d.%m.%Y %H:%M')}")
            await db_management_menu(callback)
        except Exception as e:
            await callback.message.edit_text(f"❌ Ошибка: {e}", reply_markup=get_db_management_kb())
    else:
        await callback.message.edit_text("❌ Ошибка создания бэкапа", reply_markup=get_db_management_kb())
    await callback.answer()

@router.callback_query(F.data == "db_restore_menu")
async def db_restore_menu(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("🚫 Доступ запрещен", show_alert=True); return
    all_backups = sorted(BACKUP_DIR.glob("*.db"), key=os.path.getmtime, reverse=True)
    if not all_backups:
        await callback.message.edit_text("❌ Нет бэкапов",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="⬅️ Назад", callback_data="db_management")]]))
        await callback.answer(); return
    buttons = [[InlineKeyboardButton(text=f"📅 {datetime.fromtimestamp(b.stat().st_mtime).strftime('%d.%m.%Y %H:%M')}",
                                    callback_data=f"db_restore_{b.name}")] for b in all_backups[:10]]
    buttons.append([InlineKeyboardButton(text="⬅️ Назад", callback_data="db_management")])
    await callback.message.edit_text("📥 <b>Восстановление</b>\n⚠️ Текущая БД будет заменена!",
                                    reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons))
    await callback.answer()

@router.callback_query(F.data == "db_restore_pc")
async def db_restore_pc(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("🚫 Доступ запрещен", show_alert=True); return
    await callback.answer()
    await callback.message.edit_text("📤 <b>Загрузка бэкапа с ПК</b>\n1️⃣ Нажмите 📎\n2️⃣ Выберите 'Документ'\n3️⃣ Отправьте файл .db\n⚠️ Текущая БД будет заменена!",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="⬅️ Отмена", callback_data="db_management")]]))
    await state.set_state(EditState.waiting_for_backup)
    await state.update_data(restore_mode="pc")

@router.callback_query(F.data.startswith("db_restore_"))
async def db_restore_handler(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("🚫 Доступ запрещен", show_alert=True); return
    if callback.data == "db_restore_menu":
        await db_restore_menu(callback); return
    if callback.data.startswith("db_restore_confirm_"):
        backup_name = callback.data.replace("db_restore_confirm_", "")
        backup_path = BACKUP_DIR / backup_name
        if not backup_path.exists():
            await callback.answer("❌ Файл не найден", show_alert=True); return
        await callback.message.edit_text("🔄 Восстановление...")
        try:
            current_backup = BACKUP_DIR / f"before_restore_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
            if db.db_path.exists(): shutil.copy2(db.db_path, current_backup)
            db.close()
            shutil.copy2(backup_path, db.db_path)
            db._connect()
            if db.check_integrity():
                db._execute("SELECT COUNT(*) FROM users")
                users_count = db.cursor.fetchone()[0]
                await callback.message.edit_text(f"✅ Восстановлено!\n📊 Аккаунтов: {users_count}",
                    reply_markup=InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="🗄️ Управление БД", callback_data="db_management")]]))
            else:
                if current_backup.exists():
                    db.close()
                    shutil.copy2(current_backup, db.db_path)
                    db._connect()
                await callback.message.edit_text("❌ Файл поврежден. Восстановлена предыдущая БД.",
                    reply_markup=InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="🗄️ Управление БД", callback_data="db_management")]]))
        except Exception as e:
            logger.error(f"Ошибка восстановления: {e}")
            await callback.message.edit_text(f"❌ Ошибка: {e}",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="⬅️ Назад", callback_data="db_restore_menu")]]))
        await callback.answer(); return
    # Выбор файла для подтверждения
    backup_name = callback.data.replace("db_restore_", "")
    backup_path = BACKUP_DIR / backup_name
    if not backup_path.exists():
        await callback.answer("❌ Файл не найден", show_alert=True); return
    await callback.message.edit_text(f"⚠️ <b>Подтверждение</b>\nФайл: {backup_name}\n⚠️ Текущая БД будет заменена!\nВы уверены?",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✅ Да", callback_data=f"db_restore_confirm_{backup_name}")],
            [InlineKeyboardButton(text="❌ Нет", callback_data="db_restore_menu")]]))
    await callback.answer()

@router.callback_query(F.data == "admin_export")
async def admin_export(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("🚫 Доступ запрещен", show_alert=True); return
    await callback.message.edit_text("🔄 Создание CSV...")
    path = await asyncio.to_thread(db.export_to_csv)
    if path and Path(path).exists():
        try:
            await bot.send_document(chat_id=callback.from_user.id, document=FSInputFile(path),
                                   caption=f"📤 Экспорт {datetime.now().strftime('%d.%m.%Y %H:%M')}")
            stats = db.get_stats()
            await callback.message.edit_text(f"👑 <b>Админ-панель</b>\n👥 {stats['unique_users']}", reply_markup=get_admin_kb())
        except Exception as e:
            await callback.message.edit_text(f"❌ Ошибка: {e}", reply_markup=get_admin_kb())
    else:
        await callback.message.edit_text("❌ Ошибка создания файла", reply_markup=get_admin_kb())
    await callback.answer()

@router.callback_query(F.data == "admin_export_excel")
async def admin_export_excel(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("🚫 Доступ запрещен", show_alert=True); return
    await callback.message.edit_text("🔄 Создание Excel файла...")
    path = await asyncio.to_thread(db.export_to_excel)
    if path and Path(path).exists():
        try:
            await bot.send_document(chat_id=callback.from_user.id, document=FSInputFile(path),
                                   caption=f"📊 Excel экспорт {datetime.now().strftime('%d.%m.%Y %H:%M')}")
            stats = db.get_stats()
            await callback.message.edit_text(f"👑 <b>Админ-панель</b>\n👥 {stats['unique_users']}", reply_markup=get_admin_kb())
        except Exception as e:
            await callback.message.edit_text(f"❌ Ошибка: {e}", reply_markup=get_admin_kb())
    else:
        await callback.message.edit_text("❌ Ошибка создания файла", reply_markup=get_admin_kb())
    await callback.answer()

@router.callback_query(F.data == "admin_search")
async def admin_search(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("🚫 Доступ запрещен", show_alert=True); return
    await state.set_state(EditState.waiting_search_query)
    await callback.message.edit_text("🔍 <b>Поиск</b>\nВведите ник или ID:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="❌ Отмена", callback_data="admin_back")]]))
    await callback.answer()

@router.message(EditState.waiting_search_query)
async def process_search(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id): await state.clear(); return
    query = message.text.strip()
    if len(query) < 2:
        await message.answer("❌ Минимум 2 символа"); return
    accounts = db.get_all_accounts()
    results = [acc for acc in accounts if query.lower() in acc.get('game_nickname','').lower() or query in str(acc.get('user_id',''))]
    if not results:
        await message.answer(f"❌ Ничего не найдено: {query}"); await state.clear(); return
    text = f"🔍 <b>Результаты:</b> {query}\nНайдено: {len(results)}\n" + format_accounts_table(results[:10])
    await safe_send(message, text, reply_markup=get_admin_kb())
    await state.clear()

@router.callback_query(F.data == "admin_stats")
async def admin_stats(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("🚫 Доступ запрещен", show_alert=True); return
    stats = db.get_stats()
    profile_stats = profile_db.get_stats() if profile_db else {}
    text = f"""📊 <b>Статистика</b>
👥 Пользователей: {stats['unique_users']}
🎮 Аккаунтов: {stats['total_accounts']}
📊 Профилей: {profile_stats.get('total_profiles', 0)}
✅ Активных: {profile_stats.get('active_profiles', 0)}"""
    await callback.message.edit_text(text,
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔄 Обновить", callback_data="admin_stats")],
            [InlineKeyboardButton(text="⬅️ Назад", callback_data="admin_back")]]))
    await callback.answer()

@router.callback_query(F.data == "admin_cleanup")
async def admin_cleanup(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("🚫 Доступ запрещен", show_alert=True); return
    db.cleanup_old_files(14)
    await callback.message.edit_text("🧹 <b>Очистка завершена</b>",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="🗄️ Управление БД", callback_data="db_management")]]))
    await callback.answer("✅ Готово")

@router.callback_query(F.data == "admin_refresh")
async def admin_refresh(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("🚫 Доступ запрещен", show_alert=True); return
    stats = db.get_stats()
    await callback.message.edit_text(f"👑 <b>Админ-панель</b>\n👥 {stats['unique_users']}", reply_markup=get_admin_kb())
    await callback.answer("🔄 Обновлено")

@router.callback_query(F.data == "admin_back")
async def admin_back(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("🚫 Доступ запрещен", show_alert=True); return
    stats = db.get_stats()
    await callback.message.edit_text(f"👑 <b>Админ-панель</b>\n👥 {stats['unique_users']}", reply_markup=get_admin_kb())
    await callback.answer()

@router.callback_query(F.data.startswith("admin_table_"))
async def admin_table(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("🚫 Доступ запрещен", show_alert=True); return
    try: page = int(callback.data.split("_")[2])
    except: page = 1
    accounts = db.get_all_accounts()
    if not accounts:
        await callback.message.edit_text("📋 Нет данных", reply_markup=get_admin_kb())
        await callback.answer(); return
    per_page = ACCOUNTS_PER_PAGE
    total = (len(accounts) + per_page - 1) // per_page
    page = max(1, min(page, total))
    start = (page - 1) * per_page
    end = min(start + per_page, len(accounts))
    text = f"📋 <b>Таблица</b> (стр. {page}/{total})\n" + format_accounts_table(accounts[start:end], start)
    buttons = []
    nav = []
    if page > 1: nav.append(InlineKeyboardButton(text="◀️", callback_data=f"admin_table_{page-1}"))
    nav.append(InlineKeyboardButton(text=f"{page}/{total}", callback_data="noop"))
    if page < total: nav.append(InlineKeyboardButton(text="▶️", callback_data=f"admin_table_{page+1}"))
    if nav: buttons.append(nav)
    buttons.append([InlineKeyboardButton(text="🔄 Обновить", callback_data=f"admin_table_{page}"),
                   InlineKeyboardButton(text="📤 CSV", callback_data="admin_export")])
    buttons.append([InlineKeyboardButton(text="🔍 Поиск", callback_data="admin_search"),
                   InlineKeyboardButton(text="🗑️ Пакетное удаление", callback_data="admin_batch")])
    buttons.append([InlineKeyboardButton(text="👑 В админ-панель", callback_data="admin_back")])
    await safe_send(callback, text, reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons))
    await callback.answer()

@router.callback_query(F.data == "admin_batch")
async def admin_batch(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("🚫 Доступ запрещен", show_alert=True); return
    accounts = db.get_all_accounts()
    if not accounts:
        await callback.answer("📋 Нет аккаунтов", show_alert=True); return
    await state.set_state(EditState.batch_selection)
    await state.update_data(batch_accounts=accounts, batch_selected=set(), batch_page=1)
    await show_batch_page(callback.message, state)
    await callback.answer()

async def show_batch_page(message: Message, state: FSMContext):
    data = await state.get_data()
    accounts = data.get("batch_accounts", [])
    selected = data.get("batch_selected", set())
    page = data.get("batch_page", 1)
    per_page = 10
    total_pages = (len(accounts) + per_page - 1) // per_page
    page = max(1, min(page, total_pages))
    start = (page - 1) * per_page
    end = min(start + per_page, len(accounts))
    text = f"🗑️ <b>Пакетное удаление</b> (стр. {page}/{total_pages})\nОтметьте аккаунты:\n"
    buttons = []
    for i, acc in enumerate(accounts[start:end], start + 1):
        acc_id = acc.get('id')
        nick = (acc.get('game_nickname') or '—')[:25]
        checkbox = "✅" if acc_id in selected else "⬜"
        buttons.append([InlineKeyboardButton(text=f"{checkbox} {i}. {nick}", callback_data=f"batch_toggle_{acc_id}")])
    nav = []
    if page > 1: nav.append(InlineKeyboardButton(text="◀️", callback_data="batch_page_prev"))
    nav.append(InlineKeyboardButton(text=f"{page}/{total_pages}", callback_data="noop"))
    if page < total_pages: nav.append(InlineKeyboardButton(text="▶️", callback_data="batch_page_next"))
    if nav: buttons.append(nav)
    buttons.append([InlineKeyboardButton(text="✅ Выбрать все", callback_data="batch_select_all"),
                   InlineKeyboardButton(text="⬜ Снять все", callback_data="batch_deselect_all")])
    buttons.append([InlineKeyboardButton(text="🗑️ Удалить выбранное", callback_data="batch_delete_selected"),
                   InlineKeyboardButton(text="❌ Отмена", callback_data="admin_back")])
    await message.edit_text(text, reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons))

@router.callback_query(F.data.startswith("batch_toggle_"))
async def batch_toggle(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("🚫 Доступ запрещен", show_alert=True); return
    acc_id = int(callback.data.replace("batch_toggle_", ""))
    data = await state.get_data()
    selected = data.get("batch_selected", set())
    if acc_id in selected: selected.remove(acc_id)
    else: selected.add(acc_id)
    await state.update_data(batch_selected=selected)
    await show_batch_page(callback.message, state)
    await callback.answer()

@router.callback_query(F.data.startswith("batch_select"))
async def batch_select_all(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("🚫 Доступ запрещен", show_alert=True); return
    data = await state.get_data()
    accounts = data.get("batch_accounts", [])
    selected = data.get("batch_selected", set())
    page = data.get("batch_page", 1)
    per_page = 10
    start = (page - 1) * per_page
    end = min(start + per_page, len(accounts))
    for acc in accounts[start:end]:
        if callback.data == "batch_select_all": selected.add(acc.get('id'))
        else: selected.discard(acc.get('id'))
    await state.update_data(batch_selected=selected)
    await show_batch_page(callback.message, state)
    await callback.answer()

@router.callback_query(F.data.startswith("batch_page_"))
async def batch_page_nav(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("🚫 Доступ запрещен", show_alert=True); return
    data = await state.get_data()
    page = data.get("batch_page", 1)
    if callback.data == "batch_page_next" and page < (len(data.get("batch_accounts",[])) + 9) // 10:
        await state.update_data(batch_page=page + 1)
    elif callback.data == "batch_page_prev" and page > 1:
        await state.update_data(batch_page=page - 1)
    await show_batch_page(callback.message, state)
    await callback.answer()

@router.callback_query(F.data == "batch_delete_selected")
async def batch_delete_selected(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("🚫 Доступ запрещен", show_alert=True); return
    data = await state.get_data()
    selected = data.get("batch_selected", set())
    if not selected:
        await callback.answer("❌ Нет выбранных", show_alert=True); return
    text = f"🗑️ <b>Подтверждение</b>\nВыбрано: {len(selected)}\nВы уверены?"
    await callback.message.edit_text(text,
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✅ Да", callback_data="batch_confirm_delete")],
            [InlineKeyboardButton(text="❌ Нет", callback_data="admin_batch")]]))
    await callback.answer()

@router.callback_query(F.data == "batch_confirm_delete")
async def batch_confirm_delete(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("🚫 Доступ запрещен", show_alert=True); return
    data = await state.get_data()
    selected = data.get("batch_selected", set())
    if not selected:
        await callback.answer("❌ Нет выбранных", show_alert=True); return
    deleted = []
    for acc_id in selected:
        if db.delete_account(acc_id):
            acc = db.get_account_by_id(acc_id)
            if acc: deleted.append(f"{acc['game_nickname']}")
    db.invalidate_cache()
    await state.clear()
    await callback.message.edit_text(f"✅ Удалено: {len(deleted)}\n{', '.join(deleted[:10])}",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="👑 Админ-панель", callback_data="admin_back")]]))
    await callback.answer()

@router.callback_query(F.data == "noop")
async def noop(callback: CallbackQuery):
    await callback.answer()

# ========== ФОНОВЫЕ ЗАДАЧИ ==========
async def archive_inactive_profiles():
    while True:
        try:
            await asyncio.sleep(86400)
            if not profile_db: continue
            inactive = profile_db.get_inactive_profiles(30)
            for p in inactive:
                if p.get('user_id'):
                    profile_db.archive_profile(p['user_id'])
                    logger.info(f"📦 Архивирован: {p['user_id']}")
        except Exception as e: logger.error(f"Ошибка архивации: {e}"); await asyncio.sleep(3600)

async def check_birthdays():
    while True:
        try:
            now = datetime.now()
            next_check = now.replace(hour=10, minute=0, second=0, microsecond=0)
            if now >= next_check: next_check += timedelta(days=1)
            await asyncio.sleep((next_check - now).total_seconds())
            if not profile_db: continue
            settings = profile_db.get_birthday_settings()
            if not settings: continue
            for days in [3, 1, 0]:
                profiles = profile_db.get_profiles_with_birthday_in_days(days)
                for p in profiles:
                    full_name = f"{p.get('first_name','')} {p.get('last_name','')}".strip()
                    text = f"🎉 {full_name} - День рождения!"
                    if days == 0 and settings.get('group_chat_id'):
                        try:
                            if USE_TOPIC and TARGET_TOPIC_ID:
                                await bot.send_message(chat_id=settings['group_chat_id'],
                                                      message_thread_id=TARGET_TOPIC_ID, text=text)
                            else:
                                await bot.send_message(chat_id=settings['group_chat_id'], text=text)
                        except: pass
        except Exception as e: logger.error(f"Ошибка ДР: {e}"); await asyncio.sleep(3600)

async def start_background_tasks():
    global background_tasks_started
    if background_tasks_started: return
    background_tasks_started = True
    asyncio.create_task(archive_inactive_profiles())
    asyncio.create_task(check_birthdays())
    logger.info("✅ Фоновые задачи запущены")

# ========== ЗАПУСК ==========
async def main():
    print("=" * 50)
    print("🚀 ЗАПУСК БОТА НА BOTHOST.RU")
    print("=" * 50)
    print(f"💾 БД: {db.db_path}\n👑 Админы: {ADMIN_IDS}\n🎯 Чат: {TARGET_CHAT_ID}")
    print("-" * 50)
    stats = db.get_stats()
    print(f"📊 Итог: Пользователей: {stats['unique_users']}, Аккаунтов: {stats['total_accounts']}")
    print("-" * 50)
    await start_background_tasks()
    print("📡 Режим: Polling")
    try:
        await dp.start_polling(bot)
    finally:
        db.close()

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("\n🛑 Бот остановлен")
    except Exception as e:
        print(f"\n❌ Критическая ошибка: {e}")
        traceback.print_exc()
    finally:
        try: db.close()
        except: pass
        print("👋 Завершение")
